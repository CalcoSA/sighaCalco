from app.domain.dtos.GeneratedFileDto import GeneratedFileDto
from openpyxl.utils.datetime import from_excel
from datetime import date, datetime
from openpyxl import load_workbook
from dataclasses import dataclass
from typing import Any, Callable
from pathlib import Path
from io import BytesIO
import unicodedata

@dataclass
class FieldMapping:
    targetField: str
    sourceColumns: list[str]
    transform: Callable[[dict[str, Any]], Any]

class ExcelReader:

    INCOME_DATE_COLUMN = "trabajo - fecha ingreso compania"
    COMPANY_RUT_COLUMN = "empresa - rut empresa"
    COMPANY_RUT_VALUE = "800.180.330-9"
    SALARY = "campos personalizados de trabajo - salario (obligatorio)"
    REST = "campos personalizados de trabajo - dia de descanso 2 (t)"

    def __init__(self):
        self.fieldMappings = self._buildFieldMappings()

    def generateTemplate(self, fileName: str, content: bytes, dateFrom: date, dateTo: date, previousMaster: bytes | None = None) -> GeneratedFileDto:
        sourceWorkbook = load_workbook(BytesIO(content), data_only=True, keep_links=False)
        sourceWorksheet = sourceWorkbook.active
        headerRowNumber, sourceColumnMap = self._findSourceColumns(sourceWorksheet)
        sourceRows = self._getSourceRowsByDateRange(worksheet=sourceWorksheet, headerRowNumber=headerRowNumber, columnMap=sourceColumnMap, dateFrom=dateFrom, dateTo=dateTo,)
        templatePath = self._getTemplatePath()
        templateWorkbook = load_workbook(templatePath, keep_links=False)
        templateWorksheet = templateWorkbook["Ingresos"]
        campoRow, templateFieldMap = self._findTemplateFieldColumns(worksheet=templateWorksheet, valueToFind="empleado",)

        startRow = campoRow + 1

        self._writeTransformedRows(templateWorksheet=templateWorksheet, templateFieldMap=templateFieldMap, sourceRows=sourceRows, startRow=startRow,)
        self._writeMasterChangesSheet(templateWorkbook=templateWorkbook, currentContent=content, previousContent=previousMaster)
        self._writeSalaryChangesSheet(templateWorkbook=templateWorkbook, currentContent=content, previousContent=previousMaster)
        self._writeRetirementsSheet(templateWorkbook=templateWorkbook, currentContent=content, dateFrom=dateFrom, dateTo=dateTo)
        output = BytesIO()
        templateWorkbook.save(output)
        output.seek(0)
        generatedFileName = f"Sinergy_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
        return GeneratedFileDto(fileName=generatedFileName, content=output.getvalue(), contentType="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",)

    def _buildFieldMappings(self) -> list[FieldMapping]:
        return [
            FieldMapping(
                targetField="empleado",
                sourceColumns=["colaborador - numero de documento"],
                transform=lambda row: "".join(filter(str.isdigit, self._getCellText(row["colaborador - numero de documento"]))),
            ),
            FieldMapping(
                targetField="pnombre",
                sourceColumns=["colaborador - nombre completo"],
                transform=lambda row: self._splitFullName(self._getCellText(row["colaborador - nombre completo"]))["pnombre"],
            ),
            FieldMapping(
                targetField="snombre",
                sourceColumns=["colaborador - nombre completo"],
                transform=lambda row: self._splitFullName(self._getCellText(row["colaborador - nombre completo"]))["snombre"],
            ),
            FieldMapping(
                targetField="papellido",
                sourceColumns=["colaborador - nombre completo"],
                transform=lambda row: self._splitFullName(self._getCellText(row["colaborador - nombre completo"]))["papellido"],
            ),
            FieldMapping(
                targetField="spellido",
                sourceColumns=["colaborador - nombre completo"],
                transform=lambda row: self._splitFullName(self._getCellText(row["colaborador - nombre completo"]))["spellido"],
            ),
            FieldMapping(
                targetField="fecha_nacimiento",
                sourceColumns=["colaborador - fecha de nacimiento"],
                transform=lambda row: self._formatDateAsText(row["colaborador - fecha de nacimiento"]),
            ),
            FieldMapping(
                targetField="tipo_doc_id",
                sourceColumns=["colaborador - tipo de documento"],
                transform=lambda row: self._mapDocumentTypeByCode(self._getCellText(row["colaborador - tipo de documento"])),
            ),
            FieldMapping(
                targetField="num_doc_id",
                sourceColumns=["colaborador - numero de documento"],
                transform=lambda row: "".join(filter(str.isdigit, self._getCellText(row["colaborador - numero de documento"]))),
            ),
            FieldMapping(
                targetField="ciudad_doc_id",
                sourceColumns=["campos personalizados de colaborador - lugar de expedicion- nmn"],
                transform=lambda row: self._getCellText(row["campos personalizados de colaborador - lugar de expedicion- nmn"]),
            ),
            FieldMapping(
                targetField="sexo",
                sourceColumns=["colaborador - sexo"],
                transform=lambda row: self._getCellText(row["colaborador - sexo"]),
            ),
            FieldMapping(
                targetField="fecha_ingreso",
                sourceColumns=["trabajo - fecha ingreso compania"],
                transform=lambda row: self._formatDateAsText(row["trabajo - fecha ingreso compania"]),
            ),
            FieldMapping(
                targetField="salario",
                sourceColumns=["campos personalizados de trabajo - salario (obligatorio)"],
                transform=lambda row: self._getCellText(row["campos personalizados de trabajo - salario (obligatorio)"]),
            ),
            FieldMapping(
                targetField="fecha_sueldo",
                sourceColumns=["trabajo - fecha ingreso compania"],
                transform=lambda row: self._formatDateAsText(row["trabajo - fecha ingreso compania"]),
            ),
            FieldMapping(
                targetField="tipo_sueldo",
                sourceColumns=["campos personalizados de cargo - cargo general"],
                transform=lambda row: self._mapSalaryType(self._getCellText(row["campos personalizados de cargo - cargo general"])),
            ),
            FieldMapping(
                targetField="sucursal",
                sourceColumns=[],
                transform=lambda row: "002",
            ),
            FieldMapping(
                targetField="centro_costos_1",
                sourceColumns=["trabajo - nombre sub-area asignada(o)"],
                transform=lambda row: self._mapCostCenter1(self._mapCostCenter2(self._getCellText(row["trabajo - nombre sub-area asignada(o)"]))),
            ),
            FieldMapping(
                targetField="centro_costos_2",
                sourceColumns=["trabajo - nombre sub-area asignada(o)"],
                transform=lambda row: self._mapCostCenter2(self._getCellText(row["trabajo - nombre sub-area asignada(o)"])),
            ),
            FieldMapping(
                targetField="centro_costos_3",
                sourceColumns=["trabajo - nombre sub-area asignada(o)"],
                transform=lambda row: self._mapCostCenter3(self._getCellText(row["trabajo - nombre sub-area asignada(o)"])),
            ),

            FieldMapping(
                targetField="centro_costos_4",
                sourceColumns=["campos personalizados de trabajo - dia de descanso 2 (t)"],
                transform=lambda row: self._mapRestDayCode(self._getCellText(row["campos personalizados de trabajo - dia de descanso 2 (t)"])),
            ),
            FieldMapping(
                targetField="tipo_empleado",
                sourceColumns=[],
                transform=lambda row: "002",
            ),
            FieldMapping(
                targetField="tipo_contrato",
                sourceColumns=["trabajo - cargo"],
                transform=lambda row: self._mapContractType(self._getCellText(row["trabajo - cargo"])),
            ),
            FieldMapping(
                targetField="fecha_terminacion",
                sourceColumns=["trabajo - fecha vencimiento contrato"],
                transform=lambda row: self._formatDateAsText(row["trabajo - fecha vencimiento contrato"]),
            ),
            FieldMapping(
                targetField="regimen",
                sourceColumns=[],
                transform=lambda row: "2",
            ),
            FieldMapping(
                targetField="cargo",
                sourceColumns=["trabajo - cargo"],
                transform=lambda row: self._getCellText(row["trabajo - cargo"]),
            ),
            FieldMapping(
                targetField="fondo_cesantias",
                sourceColumns=["plan - fondo de cesantia"],
                transform=lambda row: self._mapSeveranceFund(self._getCellText(row["plan - fondo de cesantia"])),
            ),
            FieldMapping(
                targetField="fecha_cesantia",
                sourceColumns=["trabajo - fecha ingreso compania"],
                transform=lambda row: self._formatDateAsText(row["trabajo - fecha ingreso compania"]),
            ),
            FieldMapping(
                targetField="entidad_pension",
                sourceColumns=["plan - fondo de pensiones"],
                transform=lambda row: self._mapPensionFund(self._getCellText(row["plan - fondo de pensiones"])),
            ),
            FieldMapping(
                targetField="sucursal_pension",
                sourceColumns=[],
                transform=lambda row: "001",
            ),
            FieldMapping(
                targetField="fecha_pension",
                sourceColumns=["trabajo - fecha ingreso compania"],
                transform=lambda row: self._formatDateAsText(row["trabajo - fecha ingreso compania"]),
            ),
            FieldMapping(
                targetField="entidad_salud",
                sourceColumns=["plan - eps"],
                transform=lambda row: self._mapEPS(self._getCellText(row["plan - eps"])),
            ),
            FieldMapping(
                targetField="sucursal_salud",
                sourceColumns=[],
                transform=lambda row: "001",
            ),
            FieldMapping(
                targetField="fecha_salud",
                sourceColumns=["trabajo - fecha ingreso compania"],
                transform=lambda row: self._formatDateAsText(row["trabajo - fecha ingreso compania"]),
            ),
            FieldMapping(
                targetField="caja_compensacion",
                sourceColumns=[],
                transform=lambda row: "CCF04",
            ),
            FieldMapping(
                targetField="fecha_caja_compensacion",
                sourceColumns=["trabajo - fecha ingreso compania"],
                transform=lambda row: self._formatDateAsText(row["trabajo - fecha ingreso compania"]),
            ),
            FieldMapping(
                targetField="corporacion",
                sourceColumns=["colaborador - banco"],
                transform=lambda row: self._mapBankCorporation(self._getCellText(row["colaborador - banco"])),
            ),
            FieldMapping(
                targetField="cuenta",
                sourceColumns=["colaborador - numero de cuenta"],
                transform=lambda row: self._getCellText(row["colaborador - numero de cuenta"]),
            ),
            FieldMapping(
                targetField="tipo_cuenta",
                sourceColumns=["colaborador - tipo de cuenta"],
                transform=lambda row: (
                    "001"
                    if self._normalize(self._getCellText(row["colaborador - tipo de cuenta"])) == "ahorro"
                    else "002"
                    if self._normalize(self._getCellText(row["colaborador - tipo de cuenta"])) == "corriente"
                    else "000"
                ),
            ),
            FieldMapping(
                targetField="sucursal_bancaria",
                sourceColumns=[],
                transform=lambda row: "001",
            ),
            FieldMapping(
                targetField="tipo_pago",
                sourceColumns=["colaborador - forma de pago"],
                transform=lambda row: (
                    "CO"
                    if self._normalize(self._getCellText(row["colaborador - forma de pago"])) == "transferencia bancaria"
                    else "CH"
                    if self._normalize(self._getCellText(row["colaborador - forma de pago"])) == "cheque"
                    else "EF"
                    if self._normalize(self._getCellText(row["colaborador - forma de pago"])) == "efectivo"
                    else ""
                ),
            ),
            FieldMapping(
                targetField="auxilio_seguro",
                sourceColumns=[],
                transform=lambda row: "4",
            ),
            FieldMapping(
                targetField="porcentaje_seguro",
                sourceColumns=[],
                transform=lambda row: "100%",
            ),
            FieldMapping(
                targetField="indicador_retención",
                sourceColumns=[],
                transform=lambda row: "2",
            ),
            FieldMapping(
                targetField="estado",
                sourceColumns=[],
                transform=lambda row: "A",
            ),
            FieldMapping(
                targetField="cuenta_gasto",
                sourceColumns=["trabajo - nombre sub-area asignada(o)"],
                transform=lambda row: self._mapCostCenter1(self._mapCostCenter2(self._getCellText(row["trabajo - nombre sub-area asignada(o)"]))),
            ),
            FieldMapping(
                targetField="entidad_riesgo",
                sourceColumns=[],
                transform=lambda row: "14-28",
            ),
            FieldMapping(
                targetField="sucur_Ent_riesgo",
                sourceColumns=[],
                transform=lambda row: "001",
            ),
            FieldMapping(
                targetField="fecha_riesgo",
                sourceColumns=["trabajo - fecha ingreso compania"],
                transform=lambda row: self._formatDateAsText(row["trabajo - fecha ingreso compania"]),
            ),
            FieldMapping(
                targetField="centro_trabajo",
                sourceColumns=["trabajo - nombre sub-area asignada(o)"],
                transform=lambda row: self._mapNewDecreeCodeByCharge(self._getCellText(row["trabajo - nombre sub-area asignada(o)"])),
            ),
            FieldMapping(
                targetField="fecha_centro_trabajo",
                sourceColumns=["trabajo - fecha ingreso compania"],
                transform=lambda row: self._formatDateAsText(row["trabajo - fecha ingreso compania"]),
            ),
            FieldMapping(
                targetField="ncontrato",
                sourceColumns=[],
                transform=lambda row: "1",
            ),
            FieldMapping(
                targetField="email",
                sourceColumns=["colaborador - email"],
                transform=lambda row: self._getCellText(row["colaborador - email"]),
            ),
            FieldMapping(
                targetField="direccion",
                sourceColumns=["colaborador - direccion"],
                transform=lambda row: self._getCellText(row["colaborador - direccion"]),
            ),
            FieldMapping(
                targetField="telefono",
                sourceColumns=["colaborador - telefono particular"],
                transform=lambda row: self._getCellText(row["colaborador - telefono particular"]),
            ),
            FieldMapping(
                targetField="estado_civil",
                sourceColumns=["colaborador - estado civil"],
                transform=lambda row: self._mapMaritalStatus(self._getCellText(row["colaborador - estado civil"])),
            ),
            FieldMapping(
                targetField="fecha_cencos",
                sourceColumns=["trabajo - fecha ingreso compania"],
                transform=lambda row: self._formatDateAsText(row["trabajo - fecha ingreso compania"]),
            ),
            FieldMapping(
                targetField="tipo_acumulado",
                sourceColumns=[],
                transform=lambda row: "N",
            ),
            FieldMapping(
                targetField="barrio",
                sourceColumns=["campos personalizados de colaborador - barrio"],
                transform=lambda row: self._getCellText(row["campos personalizados de colaborador - barrio"]),
            ),
            FieldMapping(
                targetField="pais",
                sourceColumns=["colaborador - nacionalidad"],
                transform=lambda row: (
                    "160"
                    if self._normalize(self._getCellText(row["colaborador - nacionalidad"])) == "colombiana"
                    else "000"
                ),
            ),
            FieldMapping(
                targetField="departamento",
                sourceColumns=["campos personalizados de colaborador - lugar de nacimiento"],
                transform=lambda row: self._mapDepartment(self._getCellText(row["campos personalizados de colaborador - lugar de nacimiento"])),
            ),
            FieldMapping(
                targetField="lugar_nacimiento",
                sourceColumns=["campos personalizados de colaborador - lugar de nacimiento"],
                transform=lambda row: self._mapMunicipality(self._getCellText(row["campos personalizados de colaborador - lugar de nacimiento"])),
            ),
            FieldMapping(
                targetField="tipo_sangre",
                sourceColumns=[],
                transform=lambda row: "000",
            ),
            FieldMapping(
                targetField="ojos",
                sourceColumns=[],
                transform=lambda row: "000",
            ),
            FieldMapping(
                targetField="piel",
                sourceColumns=[],
                transform=lambda row: "31",
            ),
            FieldMapping(
                targetField="cabello",
                sourceColumns=[],
                transform=lambda row: "000",
            ),
            FieldMapping(
                targetField="tipo_libreta_militar",
                sourceColumns=[],
                transform=lambda row: "N",
            ),
            FieldMapping(
                targetField="nivel_estudios",
                sourceColumns=["campos personalizados de colaborador - nivel academico"],
                transform=lambda row: self._mapAcademicLevel(self._getCellText(row["campos personalizados de colaborador - nivel academico"])),
            ),
            FieldMapping(
                targetField="empresa",
                sourceColumns=[],
                transform=lambda row: "000000000000002",
            ),
            FieldMapping(
                targetField="posicion",
                sourceColumns=["trabajo - cargo"],
                transform=lambda row: self._getCellText(row["trabajo - cargo"]),
            ),
            FieldMapping(
                targetField="tipo_cotizante",
                sourceColumns=["trabajo - cargo"],
                transform=lambda row: self._mapContributor(self._getCellText(row["trabajo - cargo"])),
            ),
            FieldMapping(
                targetField="subtipo_cotizante",
                sourceColumns=[],
                transform=lambda row: "0",
            ),
            FieldMapping(
                targetField="codigo_campo_dia_sabado",
                sourceColumns=[],
                transform=lambda row: "01",
            ),
            FieldMapping(
                targetField="trabaja_sabado",
                sourceColumns=["trabajo - cargo"],
                transform=lambda row: self._mapSaturdayWorkByCharge(self._getCellText(row["trabajo - cargo"])),
            ),
            FieldMapping(
                targetField="fecha_cambio_sabado",
                sourceColumns=["trabajo - fecha ingreso compania"],
                transform=lambda row: self._formatDateAsText(row["trabajo - fecha ingreso compania"]),
            ),
            FieldMapping(
                targetField="escalafon",
                sourceColumns=[],
                transform=lambda row: "210",
            ),
            FieldMapping(
                targetField="fecha_escalafon",
                sourceColumns=["trabajo - fecha ingreso compania"],
                transform=lambda row: self._formatDateAsText(row["trabajo - fecha ingreso compania"]),
            ),
            FieldMapping(
                targetField="auxilio_pension",
                sourceColumns=[],
                transform=lambda row: "4",
            ),
            FieldMapping(
                targetField="porcentaje_pension",
                sourceColumns=[],
                transform=lambda row: "100%",
            ),
            FieldMapping(
                targetField="auxilio_solidaridad",
                sourceColumns=[],
                transform=lambda row: "4",
            ),
            FieldMapping(
                targetField="porcentaje_solidaridad",
                sourceColumns=[],
                transform=lambda row: "100%",
            ),
            FieldMapping(
                targetField="tipo_sueldo_empleado",
                sourceColumns=["trabajo - cargo"],
                transform=lambda row: self._mapEmployeeSalaryType(self._getCellText(row["trabajo - cargo"])),
            ),
            FieldMapping(
                targetField="depto_residencia",
                sourceColumns=["colaborador - departamento"],
                transform=lambda row: self._mapDepartment(self._getCellText(row["colaborador - departamento"])),
            ),
            FieldMapping(
                targetField="municipio_resid",
                sourceColumns=["colaborador - municipio"],
                transform=lambda row: self._mapMunicipality(self._getCellText(row["colaborador - municipio"])),
            ),
        ]

    def _findSourceColumns(self, worksheet):
        requiredColumns = self._getRequiredSourceColumns()

        for row in worksheet.iter_rows(min_row=1, max_row=10):
            currentColumns = {}

            for cell in row:
                normalizedValue = self._normalize(cell.value)

                if normalizedValue in requiredColumns:
                    currentColumns[normalizedValue] = cell.column

            if all(columnName in currentColumns for columnName in requiredColumns):
                return row[0].row, currentColumns

        missingColumnsText = ", ".join(sorted(requiredColumns))

        raise ValueError(f"No se encontraron las columnas requeridas en el archivo origen: {missingColumnsText}.")

    def _getRequiredSourceColumns(self) -> set[str]:
        requiredColumns = {
            self._normalize(self.INCOME_DATE_COLUMN),
            self._normalize(self.COMPANY_RUT_COLUMN),
        }

        for mapping in self.fieldMappings:
            for sourceColumn in mapping.sourceColumns:
                requiredColumns.add(self._normalize(sourceColumn))

        return requiredColumns

    def _getSourceRowsByDateRange(self, worksheet, headerRowNumber: int, columnMap: dict, dateFrom: date, dateTo: date,) -> list[dict[str, Any]]:

        sourceRows = []

        incomeDateColumn = columnMap[self._normalize(self.INCOME_DATE_COLUMN)]
        companyRutColumn = columnMap[self._normalize(self.COMPANY_RUT_COLUMN)]
        expectedCompanyRut = self._cleanRut(self.COMPANY_RUT_VALUE)

        for rowNumber in range(headerRowNumber + 1, worksheet.max_row + 1):
            incomeDateValue = worksheet.cell(row=rowNumber, column=incomeDateColumn).value
            incomeDate = self._parseDate(incomeDateValue)

            if not incomeDate:
                continue

            if incomeDate < dateFrom or incomeDate > dateTo:
                continue

            companyRutValue = worksheet.cell(row=rowNumber, column=companyRutColumn).value
            currentCompanyRut = self._cleanRut(companyRutValue)

            if currentCompanyRut != expectedCompanyRut:
                continue

            sourceRows.append(self._buildSourceRowContext(worksheet=worksheet, rowNumber=rowNumber, columnMap=columnMap,))

        return sourceRows

    def _buildSourceRowContext(self, worksheet, rowNumber: int, columnMap: dict,) -> dict[str, Any]:
        rowContext = {}

        for columnName, columnNumber in columnMap.items():
            rowContext[columnName] = worksheet.cell(row=rowNumber, column=columnNumber,)

        return rowContext

    def _findTemplateFieldColumns(self, worksheet, valueToFind: str):
        normalizedTarget = self._normalize(valueToFind)

        for row in worksheet.iter_rows():
            for cell in row:
                if self._normalize(cell.value) == normalizedTarget:
                    campoRow = cell.row
                    templateFieldMap = self._buildTemplateFieldMap(worksheet=worksheet, campoRow=campoRow,)

                    self._validateTemplateFields(templateFieldMap)

                    return campoRow, templateFieldMap

        raise ValueError(f"No se encontró la fila {valueToFind} en la plantilla.")

    def _buildTemplateFieldMap(self, worksheet, campoRow: int) -> dict[str, int]:
        templateFieldMap = {}

        for cell in worksheet[campoRow]:
            normalizedValue = self._normalize(cell.value)

            if normalizedValue:
                templateFieldMap[normalizedValue] = cell.column

        return templateFieldMap

    def _validateTemplateFields(self, templateFieldMap: dict[str, int]) -> None:
        missingFields = []

        for mapping in self.fieldMappings:
            targetField = self._normalize(mapping.targetField)

            if targetField not in templateFieldMap:
                missingFields.append(mapping.targetField)

        if missingFields:
            raise ValueError(f"No se encontraron estos campos en la plantilla: {', '.join(missingFields)}.")

    def _writeTransformedRows(self, templateWorksheet, templateFieldMap: dict[str, int], sourceRows: list[dict[str, Any]], startRow: int,) -> None:

        for index, sourceRow in enumerate(sourceRows):
            targetRow = startRow + index

            for mapping in self.fieldMappings:
                targetField = self._normalize(mapping.targetField)
                targetColumn = templateFieldMap[targetField]
                value = mapping.transform(sourceRow)

                cell = templateWorksheet.cell(row=targetRow, column=targetColumn, value=value,)

                if targetField in {
                    "empleado",
                    "num_doc_id",
                    "fecha_nacimiento",
                    "tipo_doc_id",
                    "ciudad_doc_id",
                    "sexo",
                    "fecha_ingreso",
                    "fecha_sueldo",
                    "sucursal",
                    "centro_costos_1",
                    "centro_costos_2",
                    "centro_costos_3",
                    "tipo_empleado",
                    "fecha_terminacion",
                    "regimen",
                    "fecha_cesantia",
                    "sucursal_pension",
                    "fecha_pension",
                    "sucursal_salud",
                    "fecha_salud",
                    "caja_compensacion",
                    "fecha_caja_compensacion",
                    "corporacion",
                    "cuenta",
                    "tipo_cuenta",
                    "sucursal_bancaria",
                    "tipo_pago",
                    "auxilio_seguro",
                    "porcentaje_seguro",
                    "indicador_retención",
                    "estado",
                    "cuenta_gasto",
                    "entidad_riesgo",
                    "sucur_Ent_riesgo",
                    "fecha_riesgo",
                    "fecha_centro_trabajo",
                    "ncontrato",
                    "email",
                    "direccion",
                    "telefono",
                    "estado_civil",
                    "fecha_cencos",
                    "tipo_acumulado",
                    "barrio",
                    "tipo_sangre",
                    "ojos",
                    "piel",
                    "cabello",
                    "tipo_libreta_militar",
                    "nivel_estudios",
                    "empresa",
                    "codigo_campo_dia_sabado",
                    "fecha_cambio_sabado",
                    "escalafon",
                    "fecha_escalafon",
                    "auxilio_pension",
                    "porcentaje_pension",
                    "auxilio_solidaridad",
                    "porcentaje_solidaridad",
                }:
                    cell.number_format = "@"

    def _getCellText(self, cell) -> str:
        if cell is None:
            return ""

        value = cell.value if hasattr(cell, "value") else cell

        if value is None:
            return ""

        if isinstance(value, str):
            return value.strip()

        if isinstance(value, int):
            numberFormat = cell.number_format if hasattr(cell, "number_format") else ""
            return self._formatNumberAsText(value=value, numberFormat=numberFormat)

        if isinstance(value, float):
            if value.is_integer():
                numberFormat = cell.number_format if hasattr(cell, "number_format") else ""
                return self._formatNumberAsText(value=int(value), numberFormat=numberFormat)

            return str(value).strip()

        return str(value).strip()

    def _formatNumberAsText(self, value: int, numberFormat: str) -> str:
        normalizedFormat = str(numberFormat or "").strip()

        if normalizedFormat and set(normalizedFormat) == {"0"}:
            return str(value).zfill(len(normalizedFormat))

        return str(value)

    def _parseDate(self, value):
        if value is None:
            return None

        if isinstance(value, datetime):
            return value.date()

        if isinstance(value, date):
            return value

        if isinstance(value, (int, float)):
            try:
                return from_excel(value).date()
            except Exception:
                return None

        textValue = str(value).strip()

        for dateFormat in ("%d/%m/%Y", "%Y-%m-%d", "%d-%m-%Y"):
            try:
                return datetime.strptime(textValue, dateFormat).date()
            except ValueError:
                continue

        return None

    def _getProjectRoot(self) -> Path:
        return Path(__file__).resolve().parents[3]

    def _getTemplatePath(self) -> Path:
        templatePath = self._getProjectRoot() / "templates" / "Template_1.xlsx"

        if not templatePath.exists():
            raise ValueError(f"No se encontró la plantilla en la ruta: {templatePath}")

        return templatePath

    def _normalize(self, value) -> str:
        if value is None:
            return ""

        text = str(value).strip().lower()
        text = unicodedata.normalize("NFKD", text)
        text = "".join(character for character in text if not unicodedata.combining(character))
        text = " ".join(text.split())

        return text
    
    def _cleanRut(self, value) -> str:
        return "".join(character for character in str(value or "") if character.isdigit())
    
    def _formatDateAsText(self, value) -> str:
        if hasattr(value, "value"):
            value = value.value

        parsedDate = self._parseDate(value)

        if not parsedDate:
            return ""

        return parsedDate.strftime("%d/%m/%Y")

    def _normalizeCode(self, value: str, width: int) -> str:
        text = str(value or "").strip()

        if not text:
            return ""

        text = text.replace(",", ".")

        try:
            numericValue = float(text)

            if numericValue.is_integer():
                text = str(int(numericValue))
        except ValueError:
            pass

        if text.isdigit():
            return text.zfill(width)

        return text

    ### TRANSFORMACIONES
                    
    def _splitFullName(self, fullName: str) -> dict[str, str]:
        value = " ".join(fullName.strip().split())

        if not value:
            return {
                "pnombre": "",
                "snombre": "",
                "papellido": "",
                "spellido": "",
            }

        parts = value.split(" ")

        if len(parts) >= 4:
            papellido = parts[0]
            spellido = parts[1]
            pnombre = parts[2]
            snombre = " ".join(parts[3:])
        elif len(parts) == 3:
            papellido = parts[0]
            spellido = parts[1]
            pnombre = parts[2]
            snombre = ""
        elif len(parts) == 2:
            papellido = parts[0]
            spellido = ""
            pnombre = parts[1]
            snombre = ""
        else:
            papellido = ""
            spellido = ""
            pnombre = parts[0]
            snombre = ""

        return {
            "pnombre": pnombre.upper(),
            "snombre": snombre.upper(),
            "papellido": papellido.upper(),
            "spellido": spellido.upper(),
        }

    def _mapDocumentTypeByCode(self, value: str) -> str:
        normalizedValue = self._normalize(value)

        if not normalizedValue:
            return ""

        normalizedValue = normalizedValue.replace(",", ".")

        try:
            documentTypeCode = int(float(normalizedValue))
        except ValueError:
            documentTypeCode = None

        if documentTypeCode == 1:
            return "C"

        if documentTypeCode == 4:
            return "T"

        if documentTypeCode == 3:
            return "E"

        if documentTypeCode == 5:
            return "U"

        if "ciudadania" in normalizedValue:
            return "C"

        if "tarjeta" in normalizedValue or normalizedValue in {"t.i", "ti"}:
            return "T"

        if "extranjeria" in normalizedValue:
            return "E"

        if "permiso" in normalizedValue:
            return "U"

        return ""

    def _mapSalaryType(self, value: str) -> str:
        salaryType = str(value or "").strip()

        if not salaryType:
            return "000"

        if "aprendiz" in self._normalize(salaryType):
            return "003"

        return "001"

    def _mapCostCenter1(self, value: str) -> str:
        costCenter = str(value or "").strip()

        if not costCenter:
            return ""

        if costCenter.startswith("01"):
            return "51"

        if costCenter.startswith("02"):
            return "52"

        if costCenter.startswith("03") and costCenter.endswith("01"):
            return "73"

        if costCenter.startswith("03"):
            return "72"

        return ""

    def _mapCostCenter2(self, value: str) -> str:
        subArea = self._normalize(value)

        if not subArea:
            return ""

        subArea = subArea.replace(".", " ")
        subArea = " ".join(subArea.split())

        costCenters = {
            "amsterdam": "02005009",
            "arkadia": "02005009",
            "campestre": "02005009",
            "cocina domicilios": "02005009",
            "cocina occidente": "02005009",
            "florida etapa 2": "02005009",
            "florida parque comercial": "02005009",
            "laureles": "02005009",
            "lemont": "02005009",
            "llanogrande": "02005009",
            "mayorca": "02005009",
            "mayorca etapa dos": "02005009",
            "molinos": "02005009",
            "museo de arte moderno": "02005009",
            "one plaza": "02005009",
            "oviedo": "02005009",
            "palma grande": "02005009",
            "poblado": "02005009",
            "plaza fabricato": "02005009",
            "premium plaza": "02005009",
            "puerta del norte": "02005009",
            "punto de venta": "02005009",
            "san diego": "02005009",
            "san nicolas": "02005009",
            "santafe": "02005009",
            "tesoro": "02005009",
            "unicentro": "02005009",
            "viva envigado": "02005009",

            "h florida parque comercial": "02005009",
            "h molinos": "02005009",
            "h santafe": "02005009",
            "h tesoro": "02005009",
            "h unicentro": "02005009",
            "heladeria oviedo": "02005009",
            "heladeria viva envigado": "02005009",

            "academia de artes y formacion": "01005005",
            "almacen no perecederos": "02009003",
            "almacen perecederos": "02009004",
            "bienestar y cultura organizacional": "01005003",
            "calidad": "01011002",
            "cocina principal": "03010003",
            "compras": "01003001",
            "comunicarte": "01002008",
            "contabilidad": "01002003",
            "control y mejora continua": "01011004",
            "costos": "01002006",
            "direccion administrativa": "01004001",
            "direccion de alimentos": "03010001",
            "direccion de calidad": "01011001",
            "direccion operativa": "02008002",
            "diversidad funcional": "02005007",
            "gerencia general": "01001001",
            "gestion ambiental": "01011005",
            "gestion de activos fijos": "01004006",
            "linea bebidas": "02008007",
            "linea sal y dulce": "02008008",
            "nomina": "01002004",
            "personal volante puntos de venta": "02008004",
            "seguridad y salud en el trabajo": "01005006",
            "seleccion y contratacion": "01005004",
            "sena": "01005008",
            "servicios administrativos": "01004004",
            "servicios generales": "01004003",
            "tecnologia": "01006001",
            "tesoreria": "01002002",
            "transporte y distribucion": "02009002",
            "vinculos y relaciones humanas": "01005012",
        }

        return costCenters.get(subArea, "")

    def _mapCostCenter2_1(self, value: str) -> str:
        subArea = self._normalize(value)

        if not subArea:
            return ""

        subArea = subArea.replace(".", " ")
        subArea = " ".join(subArea.split())

        costCenters = {
            "amsterdam": "02008003",
            "arkadia": "02008003",
            "campestre": "02008003",
            "cocina domicilios": "02008003",
            "cocina occidente": "02008003",
            "florida etapa 2": "02008003",
            "florida parque comercial": "02008003",
            "laureles": "02008003",
            "lemont": "02008003",
            "llanogrande": "02008003",
            "mayorca": "02008003",
            "mayorca etapa dos": "02008003",
            "molinos": "02008003",
            "museo de arte moderno": "02008003",
            "one plaza": "02008003",
            "oviedo": "02008003",
            "palma grande": "02008003",
            "poblado": "02008003",
            "plaza fabricato": "02008003",
            "premium plaza": "02008003",
            "puerta del norte": "02008003",
            "punto de venta": "02008003",
            "san diego": "02008003",
            "san nicolas": "02008003",
            "santafe": "02008003",
            "tesoro": "02008003",
            "unicentro": "02008003",
            "viva envigado": "02008003",

            "h florida parque comercial": "02008009",
            "h molinos": "02008009",
            "h santafe": "02008009",
            "h tesoro": "02008009",
            "h unicentro": "02008009",
            "heladeria oviedo": "02008009",
            "heladeria viva envigado": "02008009",

            "academia de artes y formacion": "01005005",
            "almacen no perecederos": "02009003",
            "almacen perecederos": "02009004",
            "bienestar y cultura organizacional": "01005003",
            "calidad": "01011002",
            "cocina principal": "03010003",
            "compras": "01003001",
            "comunicarte": "01002008",
            "contabilidad": "01002003",
            "control y mejora continua": "01011004",
            "costos": "01002006",
            "direccion administrativa": "01004001",
            "direccion de alimentos": "03010001",
            "direccion de calidad": "01011001",
            "direccion operativa": "02008002",
            "diversidad funcional": "02005007",
            "gerencia general": "01001001",
            "gestion ambiental": "01011005",
            "gestion de activos fijos": "01004006",
            "linea bebidas": "02008007",
            "linea sal y dulce": "02008008",
            "nomina": "01002004",
            "personal volante puntos de venta": "02008004",
            "seguridad y salud en el trabajo": "01005006",
            "seleccion y contratacion": "01005004",
            "sena": "01005008",
            "servicios administrativos": "01004004",
            "servicios generales": "01004003",
            "tecnologia": "01006001",
            "tesoreria": "01002002",
            "transporte y distribucion": "02009002",
            "vinculos y relaciones humanas": "01005012",
        }

        return costCenters.get(subArea, "")

    def _mapCostCenter3(self, value: str) -> str:
        subArea = self._normalize(value)

        if not subArea:
            return ""

        subArea = subArea.replace(".", " ")
        subArea = " ".join(subArea.split())

        costCenter3Codes = {
            "amsterdam": "R24",
            "arkadia": "R21",
            "campestre": "R03",
            "cocina domicilios": "C01",
            "cocina occidente": "C02",
            "florida etapa 2": "R25",
            "florida parque comercial": "R15",
            "laureles": "R07",
            "lemont": "R22",
            "llanogrande": "R10",
            "mayorca": "R09",
            "mayorca etapa dos": "R17",
            "molinos": "R08",
            "museo de arte moderno": "R19",
            "one plaza": "R18",
            "oviedo": "R06",
            "palma grande": "R16",
            "poblado": "R01",
            "plaza fabricato": "R23",
            "premium plaza": "R11",
            "puerta del norte": "R14",
            "san diego": "R05",
            "san nicolas": "R13",
            "santafe": "R12",
            "tesoro": "R04",
            "unicentro": "R02",
            "viva envigado": "R20",
            "h florida parque comercial": "H05",
            "h molinos": "H04",
            "h santafe": "H02",
            "h tesoro": "H01",
            "h unicentro": "H06",
            "heladeria oviedo": "H08",
            "heladeria viva envigado": "H07",
            "academia de artes y formacion": "ADM",
            "almacen no perecederos": "LOG",
            "almacen perecederos": "LOG",
            "bienestar y cultura organizacional": "ADM",
            "calidad": "ADM",
            "cocina principal": "PPP",
            "compras": "ADM",
            "comunicarte": "ADM",
            "contabilidad": "ADM",
            "control y mejora continua": "ADM",
            "costos": "ADM",
            "direccion administrativa": "ADM",
            "direccion de alimentos": "PPP",
            "direccion de calidad": "ADM",
            "direccion operativa": "OPR",
            "diversidad funcional": "PPP",
            "gerencia general": "ADM",
            "gestion ambiental": "ADM",
            "gestion de activos fijos": "ADM",
            "linea bebidas": "TCF",
            "linea sal y dulce": "OPR",
            "nomina": "ADM",
            "personal volante puntos de venta": "MV",
            "seguridad y salud en el trabajo": "ADM",
            "seleccion y contratacion": "ADM",
            "sena": "ADM",
            "servicios administrativos": "ADM",
            "servicios generales": "ADM",
            "tecnologia": "ADM",
            "tesoreria": "ADM",
            "transporte y distribucion": "PPP",
            "vinculos y relaciones humanas": "ADM",
            "analita de datos": "ADM",
            "analitica de datos": "ADM",
            "planta de helados": "PPH",
            "planta de produccion": "PPP",
            "direccion desarrollo humano": "ADM",
            "direccion de logistica": "LOG",
            "domicilios": "OPR",
        }

        code = costCenter3Codes.get(subArea)

        if not code:
            return ""

        return f"{code}-002"

    def _mapRestDayCode(self, value: str) -> str:
        restDay = self._normalize(value)

        if not restDay:
            return ""

        restDay = restDay.replace(".", " ")
        restDay = " ".join(restDay.split())

        restDays = {
            "lunes": "001",
            "martes": "002",
            "miercoles": "003",
            "jueves": "004",
            "viernes": "005",
            "sabado": "006",
            "domingo": "007",
        }

        return restDays.get(restDay, "")

    def _mapContractType(self, typeChargeValue: str) -> str:
        typeCharge = self._normalize(typeChargeValue)

        if not typeCharge:
            return "2"

        if "aprendiz" in typeCharge and "lectiva" in typeCharge:
            return "4"

        if "aprendiz" in typeCharge and "productiva" in typeCharge:
            return "5"

        return "2"

    def _mapSeveranceFund(self, value: str) -> str:
        originalValue = " ".join(str(value or "").strip().split())

        if not originalValue:
            return "000"

        normalizedText = self._normalize(originalValue)

        if normalizedText in {"no aplica", "n/a", "na", "no"}:
            return "000"

        mappings = {
            "colfondos": "C001",
            "fna": "C002",
            "cesantias fondo nacional del ahorro": "C002",
            "porvenir": "C004",
            "proteccion": "C005",
            "skandia": "C007",
            "horizonte": "C003",
            "cesantias i n g": "C006",
            "cesantias santander": "C008",
        }

        return mappings.get(normalizedText, originalValue.upper())

    def _mapPensionFund(self, value: str) -> str:
        pensionFund = str(value or "").strip()

        if not pensionFund:
            return "P000"

        if "no aplica" in self._normalize(pensionFund):
            return "P000"

        if "colfondos" in self._normalize(pensionFund):
            return "231001"

        if "colpensiones" in self._normalize(pensionFund):
            return "25-14"

        if "colseguros" in self._normalize(pensionFund):
            return "P000"

        if "fiduciaria popular s.a." in self._normalize(pensionFund):
            return "P000"

        if "porvenir" in self._normalize(pensionFund):
            return "230301"

        if "porvenir fondo de pensiones voluntarias" in self._normalize(pensionFund):
            return "230301"

        if "proteccion" in self._normalize(pensionFund):
            return "230201"

        if "seguros sura" in self._normalize(pensionFund):
            return "P000"

        if "skandia" in self._normalize(pensionFund):
            return "230901"

        return "P000"

    def _mapEPS(self, value: str) -> str:
        eps = str(value or "").strip()

        if not eps:
            return "S000"

        if "no aplica" in self._normalize(eps):
            return "S000"

        if "aliansalud eps (antes colmedica)" in self._normalize(eps):
            return "EPS001"

        if "aliansalud" in self._normalize(eps):
            return "EPS001"

        if "asmet salud" in self._normalize(eps):
            return "S000"

        if "cafam" in self._normalize(eps):
            return "S000"

        if "capital salud" in self._normalize(eps):
            return "EPSC34"

        if "coomeva" in self._normalize(eps):
            return "EPS016"

        if "coosalud" in self._normalize(eps):
            return "ESSC24"

        if "coosalud movilidad" in self._normalize(eps):
            return "ESSC24"

        if "coosalud (subsidiado)" in self._normalize(eps):
            return "ESSC24"

        if "emssanar eps" in self._normalize(eps):
            return "ESSC18"

        if "emssanar" in self._normalize(eps):
            return "ESSC18"
        
        if "eps familiar de colombia" in self._normalize(eps):
            return "S000"

        if "famisanar" in self._normalize(eps):
            return "EPS017"

        if "famisanar l.t.d.a." in self._normalize(eps):
            return "EPS017"
        
        if "fosyga" in self._normalize(eps):
            return "MIN002"

        if "e.p.s. fosyga" in self._normalize(eps):
            return "MIN001"

        if "fundación salud mia eps" in self._normalize(eps):
            return "S000"

        if "mutual ser" in self._normalize(eps):
            return "S000"

        if "mutual ser (subsidiado)" in self._normalize(eps):
            return "S000"

        if "nueva e.p.s." in self._normalize(eps):
            return "EPS037"

        if "nueva eps movilidad" in self._normalize(eps):
            return "S000"

        if "proteger eps s.a.s" in self._normalize(eps):
            return "S000"

        if "s.o.s. servicio occidental de salud s.a." in self._normalize(eps):
            return "EPS018"
        
        if "servicio occidental de salud s.o.s." in self._normalize(eps):
            return "EPS018"

        if "savia salud" in self._normalize(eps):
            return "S000"     

        if "salud total" in self._normalize(eps):
            return "EPS002"

        if "sanitas" in self._normalize(eps):
            return "EPS005"

        if "compensar" in self._normalize(eps):
            return "EPS008"

        if "eps sura" in self._normalize(eps):
            return "EPS010"

        if "sura" in self._normalize(eps):
            return "EPS010"

        if "comfenalco valle" in self._normalize(eps):
            return "EPS012"

        if "comfamiliar del huila" in self._normalize(eps):
            return "CCFC24"

        if "comfaoriente" in self._normalize(eps):
            return "CCFC50"

        if "comfacundi" in self._normalize(eps):
            return "CCFC53"

        if "caja de compensacion familiar de cundinamarca" in self._normalize(eps):
            return "CCFC53"

        if "cajacopi" in self._normalize(eps):
            return "CCFC55"

        if "cafesalud" in self._normalize(eps):
            return "EPS003"

        if "instituto de seguros sociales" in self._normalize(eps):
            return "EPS006"

        if "comfenalco antioquia" in self._normalize(eps):
            return "EPS009"

        if "saludcoop" in self._normalize(eps):
            return "EPS013"

        if "humana vivir" in self._normalize(eps):
            return "EPS014"

        if "salud colpatria" in self._normalize(eps):
            return "EPS015"

        if "colpatria" in self._normalize(eps):
            return "EPS015"

        if "caprecom" in self._normalize(eps):
            return "EPS020"

        if "cruz blanca" in self._normalize(eps):
            return "EPS023"

        if "cajanal" in self._normalize(eps):
            return "EPS024"

        if "capresoca" in self._normalize(eps):
            return "EPS025"

        if "sol salud" in self._normalize(eps):
            return "EPS026"

        if "salud vida" in self._normalize(eps):
            return "EPS033"

        if "red salud" in self._normalize(eps):
            return "EPS035"

        if "nueva s a" in self._normalize(eps):
            return "EPS037"

        if "golden group" in self._normalize(eps):
            return "EPS039"

        if "medimas" in self._normalize(eps):
            return "EPS044"

        if "asociacion mutual ser empresa solidaria de salud" in self._normalize(eps):
            return "EPSC07"

        if "asociacion mutual ser empresa solidaria de salus" in self._normalize(eps):
            return "EPSC07"

        if "convida" in self._normalize(eps):
            return "EPSC22"

        if "asociacion indigena del cauca" in self._normalize(eps):
            return "EPSIC3"

        if "eps comparte" in self._normalize(eps):
            return "ESS133"

        if "mutual ser empresa solidaria de salud" in self._normalize(eps):
            return "ESSC07"

        if "comparta" in self._normalize(eps):
            return "ESSC33"

        return "S000"

    def _mapBankCorporation(self, value: str) -> str:
        bankCorporation = str(value or "").strip()

        if not bankCorporation:
            return ""

        if "no aplica" in self._normalize(bankCorporation):
            return ""

        if "banco av villas" in self._normalize(bankCorporation):
            return "107"

        if "banco caja social bcsc" in self._normalize(bankCorporation):
            return "330"

        if "banco caja social" in self._normalize(bankCorporation):
            return "330"

        if "banco davivienda" in self._normalize(bankCorporation):
            return "104"

        if "bancolombia" in self._normalize(bankCorporation):
            return "303"

        if "bbva colombia" in self._normalize(bankCorporation):
            return "241"

        if "bbva" in self._normalize(bankCorporation):
            return "241"

        if "davibank" in self._normalize(bankCorporation):
            return ""

        if "sin definir" in self._normalize(bankCorporation):
            return "000"

        if "scotia bank" in self._normalize(bankCorporation):
            return "101"

        if "colpatria" in self._normalize(bankCorporation):
            return "101"

        if "colmena" in self._normalize(bankCorporation):
            return "108"

        if "banco de bogota" in self._normalize(bankCorporation):
            return "301"

        if "banco popular" in self._normalize(bankCorporation):
            return "302"

        if "citibank" in self._normalize(bankCorporation):
            return "309"

        if "banco agrario" in self._normalize(bankCorporation):
            return "311"

        if "banco sudameris colombia" in self._normalize(bankCorporation):
            return "312"

        if "sudameris" in self._normalize(bankCorporation):
            return "312"

        if "helm bank" in self._normalize(bankCorporation):
            return "314"

        if "banco de occidente" in self._normalize(bankCorporation):
            return "323"

        if "banco santander" in self._normalize(bankCorporation):
            return "339"

        if "santander" in self._normalize(bankCorporation):
            return "339"

        if "hsbc" in self._normalize(bankCorporation):
            return "345"

        if "banco falabella" in self._normalize(bankCorporation):
            return "346"

        if "falabella" in self._normalize(bankCorporation):
            return "346"

        if "coomeva banco" in self._normalize(bankCorporation):
            return "347"

        if "banco coomeva" in self._normalize(bankCorporation):
            return "347"

        if "itau" in self._normalize(bankCorporation):
            return "348"

        if "corpbanca" in self._normalize(bankCorporation):
            return "348"

        if "gnb" in self._normalize(bankCorporation):
            return "349"

        return ""

    def _mapNewDecreeCodeByCharge(self, value: str) -> str:
        subArea = self._normalize(value)

        if not subArea:
            return ""

        subArea = subArea.replace(".", " ")
        subArea = " ".join(subArea.split())

        newDecreeCodes = {
            "amsterdam": "3561101",
            "arkadia": "3561101",
            "campestre": "3561101",
            "cocina domicilios": "3561101",
            "cocina occidente": "3561101",
            "florida etapa 2": "3561101",
            "florida parque comercial": "3561101",
            "laureles": "3561101",
            "lemont": "3561101",
            "llanogrande": "3561101",
            "mayorca": "3561101",
            "mayorca etapa dos": "3561101",
            "molinos": "3561101",
            "museo de arte moderno": "3561101",
            "one plaza": "3561101",
            "oviedo": "3561101",
            "palma grande": "3561101",
            "poblado": "3561101",
            "plaza fabricato": "3561101",
            "premium plaza": "3561101",
            "puerta del norte": "3561101",
            "punto de venta": "3561101",
            "san diego": "3561101",
            "san nicolas": "3561101",
            "santafe": "3561101",
            "tesoro": "3561101",
            "unicentro": "3561101",
            "viva envigado": "3561101",
            "h florida parque comercial": "3561101",
            "h molinos": "3561101",
            "h santafe": "3561101",
            "h tesoro": "3561101",
            "h unicentro": "3561101",
            "heladeria oviedo": "3561101",
            "heladeria viva envigado": "3561101",
            "academia de artes y formacion": "1701001",
            "almacen no perecederos": "2521001",
            "almacen perecederos": "2521001",
            "bienestar y cultura organizacional": "1701001",
            "calidad": "1701001",
            "cocina principal": "2108901",
            "compras": "1701001",
            "comunicarte": "1701001",
            "contabilidad": "1701001",
            "control y mejora continua": "1701001",
            "costos": "1701001",
            "direccion administrativa": "1701001",
            "direccion de alimentos": "2108901",
            "direccion de calidad": "1701001",
            "direccion operativa": "2108901",
            "diversidad funcional": "2108901",
            "gerencia general": "1701001",
            "gestion ambiental": "1701001",
            "gestion de activos fijos": "1701001",
            "linea bebidas": "3561101",
            "linea sal y dulce": "2108901",
            "nomina": "1701001",
            "personal volante puntos de venta": "3561101",
            "seguridad y salud en el trabajo": "1701001",
            "seleccion y contratacion": "1701001",
            "sena": "1701001",
            "servicios administrativos": "1701001",
            "servicios generales": "1701001",
            "tecnologia": "1701001",
            "tesoreria": "1701001",
            "transporte y distribucion": "2108901",
            "vinculos y relaciones humanas": "1701001",
            "analita de datos": "1701001",
            "analitica de datos": "1701001",
            "planta de helados": "2108901",
            "planta de produccion": "2108901",
            "direccion desarrollo humano": "1701001",
            "direccion de logistica": "2521001",
            "domicilios": "2108901",
        }

        return newDecreeCodes.get(subArea, "")

    def _mapMaritalStatus(self, value: str) -> str:
        maritalStatus = self._normalize(value)

        if not maritalStatus:
            return "000"

        if maritalStatus == "soltero":
            return "001"

        if maritalStatus == "casado":
            return "002"

        if maritalStatus == "union libre":
            return "003"

        if maritalStatus == "viudo":
            return "004"

        if maritalStatus == "divorciado":
            return "005"

        if maritalStatus == "separado":
            return "005"

        return "000"

    def _mapDepartment(self, value: str) -> str:
        municipalityCode, departmentCode = self._resolveBirthPlace(value)
        return departmentCode

    def _mapMunicipality(self, value: str) -> str:
        municipalityCode, departmentCode = self._resolveBirthPlace(value)
        return municipalityCode

    def _resolveBirthPlace(self, value: str) -> tuple[str, str]:
        place = self._normalize(value)

        if not place:
            return "000", "000"

        place = (
            place.replace(".", " ")
            .replace(",", " ")
            .replace("-", " ")
            .replace("/", " ")
            .replace("_", " ")
            .replace("(", " ")
            .replace(")", " ")
        )

        place = " ".join(place.split())

        if not place:
            return "000", "000"

        foreignKeywords = [
            "venezuela",
            "venezuala",
            "caracas",
            "maracaibo",
            "merida",
            "tachira",
            "zulia",
            "aragua",
            "guarico",
            "lara",
            "barquisimeto",
            "maracay",
            "ciudad ojeda",
            "portuguesa",
            "acarigua",
            "barinas",
            "san cristobal",
            "ecuador",
            "suiza",
            "zurich",
        ]

        if any(keyword in place for keyword in foreignKeywords):
            return "1125", "000"

        departmentCatalog = {
            "sin definir": "000",
            "antioquia": "05",
            "atlantico": "08",
            "atlántico": "08",
            "bogota": "11",
            "bogota dc": "11",
            "bogota d c": "11",
            "bogota d.c": "11",
            "bolivar": "13",
            "boyaca": "15",
            "caldas": "17",
            "caqueta": "18",
            "cauca": "19",
            "cesar": "20",
            "cordoba": "23",
            "cundinamarca": "25",
            "choco": "27",
            "huila": "41",
            "la guajira": "44",
            "guajira": "44",
            "magdalena": "47",
            "meta": "50",
            "narino": "52",
            "nariño": "52",
            "n santander": "54",
            "norte de santander": "54",
            "quindio": "63",
            "risaralda": "66",
            "santander": "68",
            "sucre": "70",
            "tolima": "73",
            "valle": "76",
            "valle del cauca": "76",
            "valle del cauc": "76",
            "arauca": "81",
            "casanare": "85",
            "putumayo": "86",
            "san andres": "88",
            "amazonas": "91",
            "guainia": "94",
            "guaviare": "95",
            "vaupes": "97",
            "vichada": "99",
        }

        if place in departmentCatalog:
            return "000", departmentCatalog[place]

        bogotaKeywords = [
            "bogota",
            "bogota dc",
            "bogota d c",
            "bogota d c",
            "bogota d.c",
            "bogota d.c.",
            "suba",
            "bosa",
            "kennedy",
            "usme",
            "usaquen",
            "engativa",
            "fontibon",
            "tunjuelito",
            "chapinero",
            "puente aranda",
            "ciudad bolivar",
            "rafael uribe",
            "san cristobal sur",
            "antonio narino",
        ]

        if any(keyword in place for keyword in bogotaKeywords):
            return "167", "11"

        aliases = {
            "cartagena de indias": "cartagena",
            "cartagena bolivar": "cartagena",
            "cartagena bolivar": "cartagena",
            "monteria cordoba": "monteria",
            "monteria cordoba": "monteria",
            "quibdo choco": "quibdo",
            "quibdo choco": "quibdo",
            "istmina": "itsmina",
            "istmina choco": "itsmina",
            "itsmina choco": "itsmina",
            "novita choco": "novita",
            "jovita choco": "novita",
            "mompox": "mompos",
            "monpox": "mompos",
            "mompos": "mompos",
            "tolu": "santiago de tolu",
            "tolu viejo": "tolu viejo",
            "since sucre": "since",
            "sincelejo sucre": "sincelejo",
            "sanonofre sucre": "san onofre",
            "san onofre sucre": "san onofre",
            "san pedro sucre": "san pedro",
            "sam pedro de uraba": "san pedro de uraba",
            "san pedro de uraba": "san pedro de uraba",
            "rio negro": "rionegro",
            "rionegro santander": "rionegro",
            "cucuta norte de santander": "cucuta",
            "cucuta n santander": "cucuta",
            "cucuta norte de stander": "cucuta",
            "cucuta n/santander": "cucuta",
            "canas gordas": "canasgordas",
            "canasgordas": "canasgordas",
            "barancabermeja": "barrancabermeja",
            "vistahermosa": "vista hermosa",
            "vista hermosa meta": "vista hermosa",
            "san vicente del caguan caqueta": "san vicente del caguan",
            "florencia caqueta": "florencia",
            "la dorada caldas": "la dorada",
            "manizales caldas": "manizales",
            "chinchina caldas": "chinchina",
            "supia caldas": "supia",
            "palmira valle": "palmira",
            "palmira valle del cauca": "palmira",
            "cali valle": "cali",
            "cali valle del cauca": "cali",
            "santiago de cali": "cali",
            "cali colombia": "cali",
            "yumbo valle": "yumbo",
            "zarzal valle": "zarzal",
            "florida valle": "florida",
            "tulua valle del cauca": "tulua",
            "buga valle": "guadalajara de buga",
            "buenaventura valle del cauca": "buenaventura",
            "tumaco narino": "tumaco",
            "barbacoa narino": "barbacoas",
            "barbacoas narino": "barbacoas",
            "magui narino": "magui",
            "satinga narino": "olaya herrera",
            "olaya herrera narino": "olaya herrera",
            "el charco narino": "el charco",
            "charco narino": "el charco",
            "pasto narino": "pasto",
            "ipiales narino": "ipiales",
            "popayan cauca": "popayan",
            "popayan cauca": "popayan",
            "el tambo cauca": "el tambo",
            "tambo cauca": "el tambo",
            "patia el bordo": "patia",
            "patia el bordo cauca": "patia",
            "bordo cauca": "patia",
            "el bordo cauca": "patia",
            "miranda cauca": "miranda",
            "guapi cauca": "guapi",
            "timbiqui cauca": "timbiqui",
            "paez cauca": "paez",
            "inza cauca": "inza",
            "santander de quilichao cauca": "santander de quilichao",
            "riohacha la guajira": "riohacha",
            "rio hacha la guajira": "riohacha",
            "maicao guajira": "maicao",
            "uribia guajira": "uribia",
            "fonseca guajira": "fonseca",
            "valledupar cesar": "valledupar",
            "valledupar cesar": "valledupar",
            "curumani cesar": "curumani",
            "chimichagua cesar": "chimichagua",
            "agustin codazzi": "agustin codazzi",
            "la paz cesar": "la paz",
            "astrea cesar": "astrea",
            "bosconia cesar": "bosconia",
            "santa marta magdalena": "santa marta",
            "fundacion magdalena": "fundacion",
            "fundacion magdalena": "fundacion",
            "el banco magdalena": "el banco",
            "cienaga magdalena": "cienaga",
            "ariguani magdalena": "ariguani",
            "guamal magdalena": "guamal",
            "pivijay magdalena": "pivijay",
            "nueva granada magdalena": "nueva granada",
            "san sebastian de buenavista magdalena": "san sebastian de buenavista",
            "barranquilla atlantico": "barranquilla",
            "soledad atlantico": "soledad",
            "malambo atlantico": "malambo",
            "luruaco atlantico": "luruaco",
            "santo tomas atlantico": "santo tomas",
            "suan atlantico": "suan",
            "baranoa atlantico": "baranoa",
            "moniquira boyaca": "moniquira",
            "moniquira boyaca": "moniquira",
            "tunja boyaca": "tunja",
            "tunja boyaca": "tunja",
            "sogamoso boyaca": "sogamoso",
            "duitama boyaca": "duitama",
            "muzo boyaca": "muzo",
            "chiquinquira boyaca": "chiquinquira",
            "paipa boyaca": "paipa",
            "ramiriqui boyaca": "ramiriqui",
            "turmeque boyaca": "turmeque",
            "guateque boyaca": "guateque",
            "sotaquira boyaca": "sotaquira",
            "mogui boyaca": "mongui",
            "mongui boyaca": "mongui",
            "socorro santander": "socorro",
            "bucaramanga santander": "bucaramanga",
            "barrancabermeja santander": "barrancabermeja",
            "cimitarra santander": "cimitarra",
            "floridablanca santander": "floridablanca",
            "charala santander": "charala",
            "mogotes santander": "mogotes",
            "puente nacional santander": "puente nacional",
            "malaga santander": "malaga",
            "barbosa santander": "barbosa",
            "sabana de torres santander": "sabana de torres",
            "ocamonte santander": "ocamonte",
            "piedecuesta santander": "piedecuesta",
            "la paz santander": "la paz",
            "ibague tolima": "ibague",
            "ibague tolima": "ibague",
            "melgar tolima": "melgar",
            "fresno tolima": "fresno",
            "honda tolima": "honda",
            "planadas tolima": "planadas",
            "chaparral tolima": "chaparral",
            "icononzo tolima": "icononzo",
            "ortega tolima": "ortega",
            "purificacion tolima": "purificacion",
            "coyaima tolima": "coyaima",
            "flandes tolima": "flandes",
            "mariquita tolima": "mariquita",
            "garzon huila": "garzon",
            "neiva huila": "neiva",
            "pitalito huila": "pitalito",
            "la plata huila": "la plata",
            "algeciras huila": "algeciras",
            "tarqui huila": "tarqui",
            "villavicencio meta": "villavicencio",
            "acacias meta": "acacias",
            "puerto lopez meta": "puerto lopez",
            "lejanias meta": "lejanias",
            "mapiripan meta": "mapiripan",
            "el castillo meta": "el castillo",
            "san juan de arama": "san juan de arama",
            "yopal casanare": "yopal",
            "tauramena casanare": "tauramena",
            "villanueva casanare": "villanueva",
            "mocoa putumayo": "mocoa",
            "orito putumayo": "orito",
            "arauca arauca": "arauca",
            "tame arauca": "tame",
            "san jose del guaviare": "san jose del guaviare",
            "puerto narino": "puerto narino",
        }

        for alias, realPlace in aliases.items():
            if alias == place or alias in place:
                place = realPlace
                break

        detectedDepartment = ""

        for departmentName, departmentCode in sorted(departmentCatalog.items(), key=lambda item: len(item[0]), reverse=True,):
            if departmentName in place and departmentCode != "000":
                detectedDepartment = departmentCode
                break

        municipalities = {
            "abejorral": [("013", "05")],
            "abriaqui": [("014", "05")],
            "acacias": [("735", "50")],
            "acandi": [("465", "27")],
            "achi": [("169", "13")],
            "agua de dios": [("523", "25")],
            "aguachica": [("440", "20")],
            "aguadas": [("337", "17")],
            "agustin codazzi": [("441", "20")],
            "algeciras": [("656", "41")],
            "alejandria": [("015", "05")],
            "alpujarra": [("1022", "73")],
            "amalfi": [("017", "05")],
            "amaga": [("016", "05")],
            "andes": [("018", "05")],
            "angostura": [("020", "05")],
            "anori": [("021", "05")],
            "anserma": [("338", "17")],
            "ansermanuevo": [("1071", "76")],
            "apartado": [("024", "05")],
            "aranzazu": [("339", "17")],
            "araucua": [("137", "81")],
            "arauca": [("137", "81")],
            "arbelaez": [("527", "25")],
            "arboletes": [("025", "05")],
            "arenal": [("171", "13")],
            "argelia": [("026", "05"), ("400", "19"), ("1072", "76")],
            "ariguani": [("707", "47")],
            "arjona": [("172", "13")],
            "armenia": [("880", "63"), ("027", "05")],
            "astrea": [("442", "20")],
            "ayapel": [("496", "23")],
            "bagado": [("468", "27")],
            "bahia solano": [("469", "27")],
            "bajo baudo": [("470", "27")],
            "balboa": [("401", "19"), ("894", "66")],
            "baranoa": [("145", "08")],
            "barbacoas": [("768", "52")],
            "barbosa": [("028", "05"), ("912", "68")],
            "barrancabermeja": [("914", "68")],
            "barranquilla": [("144", "08")],
            "becerril": [("443", "20")],
            "belalcazar": [("340", "17")],
            "belen de umbria": [("895", "66")],
            "bello": [("030", "05")],
            "betulia": [("032", "05"), ("915", "68")],
            "bogota": [("167", "11")],
            "bogota d c": [("167", "11")],
            "bogota dc": [("167", "11")],
            "bogota d.c": [("167", "11")],
            "bojaya": [("472", "27")],
            "bosconia": [("444", "20")],
            "bucaramanga": [("908", "68")],
            "buenaventura": [("1074", "76")],
            "buenavista": [("497", "23"), ("881", "63"), ("996", "70"), ("223", "15")],
            "buga": [("1075", "76")],
            "caceres": [("036", "05")],
            "cachipay": [("532", "25")],
            "caicedo": [("037", "05")],
            "caimito": [("997", "70")],
            "cajica": [("533", "25")],
            "calarca": [("882", "63")],
            "caldas": [("038", "05"), ("225", "15")],
            "cali": [("1068", "76")],
            "caloto": [("406", "19")],
            "campamento": [("039", "05")],
            "canalete": [("498", "23")],
            "canasgordas": [("040", "05")],
            "cañasgordas": [("040", "05")],
            "caramanta": [("042", "05")],
            "carepa": [("043", "05")],
            "carmen de viboral": [("044", "05")],
            "cartagena": [("168", "13")],
            "cartagena del chaira": [("366", "18")],
            "cartago": [("1080", "76")],
            "caucasia": [("046", "05")],
            "cerete": [("499", "23")],
            "cerrito": [("922", "68"), ("1084", "76")],
            "cerro san antonio": [("708", "47")],
            "chaparral": [("1031", "73")],
            "charala": [("923", "68")],
            "chia": [("538", "25")],
            "chigorodo": [("047", "05")],
            "chimichagua": [("445", "20")],
            "chinacota": [("834", "54")],
            "chinchina": [("341", "17")],
            "chinu": [("501", "23")],
            "chiquinquira": [("229", "15")],
            "chitaraque": [("232", "15")],
            "chivolo": [("709", "47")],
            "choachi": [("540", "25")],
            "choconta": [("541", "25")],
            "cienaga": [("710", "47")],
            "cimitarra": [("927", "68")],
            "ciudad bolivar": [("033", "05")],
            "cocorna": [("049", "05")],
            "coloso": [("998", "70")],
            "combita": [("235", "15")],
            "concordia": [("051", "05"), ("711", "47")],
            "condoto": [("476", "27")],
            "copacabana": [("052", "05")],
            "corozal": [("999", "70")],
            "cota": [("543", "25")],
            "coyaima": [("1033", "73")],
            "cucuta": [("827", "54")],
            "curumani": [("447", "20")],
            "dabeiba": [("053", "05")],
            "dos quebradas": [("896", "66")],
            "dosquebradas": [("896", "66")],
            "duitama": [("244", "15")],
            "el bagre": [("056", "05")],
            "el banco": [("712", "47")],
            "el bordo": [("421", "19")],
            "el castillo": [("742", "50")],
            "el cerrito": [("1084", "76")],
            "el charco": [("779", "52")],
            "el tambo": [("408", "19"), ("783", "52")],
            "envigado": [("058", "05")],
            "espinal": [("1036", "73")],
            "facatativa": [("548", "25")],
            "filadelfia": [("342", "17")],
            "flandes": [("1038", "73")],
            "florencia": [("363", "18"), ("409", "19")],
            "florida": [("1086", "76")],
            "floridablanca": [("940", "68")],
            "fonseca": [("695", "44")],
            "fortul": [("140", "81")],
            "fredonia": [("059", "05")],
            "fresno": [("1039", "73")],
            "frontino": [("060", "05")],
            "fundacion": [("715", "47")],
            "fusagasuga": [("553", "25")],
            "gacheta": [("556", "25")],
            "gameza": [("250", "15")],
            "garzon": [("662", "41")],
            "giraldo": [("061", "05")],
            "girardot": [("558", "25")],
            "girardota": [("062", "05")],
            "giron": [("943", "68")],
            "gomez plata": [("063", "05")],
            "granada": [("064", "05"), ("559", "25"), ("745", "50")],
            "guadalajara de buga": [("1075", "76")],
            "guaduas": [("561", "25")],
            "guamal": [("716", "47"), ("746", "50")],
            "guamo": [("1040", "73")],
            "guapi": [("410", "19")],
            "guaranda": [("1004", "70")],
            "guasca": [("562", "25")],
            "guateque": [("253", "15")],
            "guatica": [("897", "66")],
            "honda": [("1042", "73")],
            "ibague": [("1021", "73")],
            "icononzo": [("1043", "73")],
            "inza": [("411", "19")],
            "ipiales": [("790", "52")],
            "itagui": [("070", "05")],
            "itsmina": [("479", "27")],
            "istmina": [("479", "27")],
            "ituango": [("071", "05")],
            "iza": [("256", "15")],
            "jamundi": [("1089", "76")],
            "jardin": [("072", "05")],
            "jerico": [("073", "05"), ("258", "15")],
            "junin": [("569", "25")],
            "la ceja": [("074", "05")],
            "la dorada": [("343", "17")],
            "la mesa": [("571", "25")],
            "la paz": [("459", "20"), ("954", "68")],
            "la plata": [("669", "41")],
            "la sierra": [("413", "19")],
            "la union": [("077", "05"), ("1005", "70"), ("1091", "76"), ("795", "52")],
            "la vega": [("414", "19"), ("574", "25")],
            "lenguazaque": [("575", "25")],
            "lerida": [("1044", "73")],
            "libano": [("1045", "73")],
            "liborina": [("078", "05")],
            "lloro": [("481", "27")],
            "lorica": [("505", "23")],
            "los cordobas": [("506", "23")],
            "los palmitos": [("1006", "70")],
            "luruaco": [("150", "08")],
            "machica": [("576", "25")],
            "madrid": [("577", "25")],
            "magangue": [("184", "13")],
            "magui": [("799", "52")],
            "mahates": [("185", "13")],
            "maicao": [("698", "44")],
            "majagual": [("1007", "70")],
            "malaga": [("958", "68")],
            "malambo": [("151", "08")],
            "manizales": [("336", "17")],
            "mapiripan": [("747", "50")],
            "maria la baja": [("187", "13")],
            "marinilla": [("080", "05")],
            "mariquita": [("1046", "73")],
            "marsella": [("900", "66")],
            "medellin": [("012", "05")],
            "melgar": [("1047", "73")],
            "mercaderes": [("416", "19")],
            "miranda": [("417", "19")],
            "mocoa": [("867", "86")],
            "mogotes": [("960", "68")],
            "mompox": [("189", "13")],
            "mompos": [("189", "13")],
            "mongui": [("268", "15")],
            "moniquira": [("269", "15")],
            "montebello": [("081", "05")],
            "montelibano": [("508", "23")],
            "montenegro": [("888", "63")],
            "monteria": [("495", "23")],
            "moñitos": [("509", "23")],
            "morales": [("190", "13"), ("418", "19")],
            "morroa": [("1008", "70")],
            "murindo": [("082", "05")],
            "mutata": [("083", "05")],
            "muzo": [("271", "15")],
            "narino": [("084", "05"), ("581", "25"), ("802", "52")],
            "nechi": [("086", "05")],
            "necocli": [("085", "05")],
            "neiva": [("652", "41")],
            "novita": [("485", "27")],
            "nueva granada": [("717", "47")],
            "nuqui": [("486", "27")],
            "obando": [("1093", "76")],
            "ocamonte": [("962", "68")],
            "olaya herrera": [("803", "52")],
            "orito": [("869", "86")],
            "ortega": [("1050", "73")],
            "otanche": [("275", "15")],
            "ovejas": [("1009", "70")],
            "paez": [("277", "15"), ("420", "19")],
            "paime": [("588", "25")],
            "paipa": [("278", "15")],
            "palestina": [("352", "17"), ("674", "41")],
            "palmira": [("1094", "76")],
            "pamplona": [("852", "54")],
            "pasca": [("591", "25")],
            "pasto": [("763", "52")],
            "patia": [("421", "19")],
            "peque": [("089", "05")],
            "pereira": [("892", "66")],
            "piedecuesta": [("968", "68")],
            "pitalito": [("676", "41")],
            "pivijay": [("720", "47")],
            "planadas": [("1053", "73")],
            "planeta rica": [("510", "23")],
            "plato": [("721", "47")],
            "popayan": [("398", "19")],
            "pueblo nuevo": [("511", "23")],
            "pueblorrico": [("090", "05")],
            "pueblo viejo": [("722", "47")],
            "puente nacional": [("970", "68")],
            "puerto berrio": [("091", "05")],
            "puerto boyaca": [("286", "15")],
            "puerto escondido": [("512", "23")],
            "puerto libertador": [("513", "23")],
            "puerto lopez": [("754", "50")],
            "puerto narino": [("009", "91")],
            "puerto salgar": [("592", "25")],
            "puerto triunfo": [("093", "05")],
            "purificacion": [("1055", "73")],
            "purisima": [("514", "23")],
            "quibdo": [("464", "27")],
            "quinchia": [("903", "66")],
            "quipama": [("287", "15")],
            "ramiriqui": [("288", "15")],
            "retiro": [("095", "05")],
            "riohacha": [("689", "44")],
            "rionegro": [("096", "05"), ("973", "68")],
            "riosucio": [("354", "17"), ("489", "27")],
            "risaralda": [("355", "17")],
            "rosas": [("426", "19")],
            "sabana de torres": [("974", "68")],
            "sabanalarga": [("097", "05"), ("160", "08"), ("391", "85")],
            "sabaneta": [("098", "05")],
            "saboya": [("291", "15")],
            "sahagun": [("515", "23")],
            "salgar": [("099", "05")],
            "samana": [("357", "17")],
            "san alberto": [("460", "20")],
            "san andres": [("100", "05"), ("906", "88"), ("975", "68")],
            "san antonio": [("1060", "73")],
            "san benito abad": [("1012", "70")],
            "san bernardo": [("600", "25"), ("815", "52")],
            "san bernardo del viento": [("518", "23")],
            "san carlos": [("101", "05"), ("519", "23")],
            "san diego": [("461", "20")],
            "san estanislao": [("195", "13")],
            "san fernando": [("196", "13")],
            "san francisco": [("102", "05"), ("602", "25"), ("875", "86")],
            "san jacinto": [("197", "13")],
            "san jacinto del cauca": [("198", "13")],
            "san jeronimo": [("103", "05")],
            "san jose de la fragua": [("374", "18")],
            "san jose de la montaña": [("104", "05")],
            "san jose del guaviare": [("648", "95")],
            "san jose del palmar": [("490", "27")],
            "san juan de arama": [("759", "50")],
            "san juan de uraba": [("105", "05")],
            "san juan nepomuceno": [("199", "13")],
            "san luis": [("106", "05"), ("1061", "73")],
            "san marcos": [("1014", "70")],
            "san martin": [("462", "20"), ("761", "50")],
            "san martin de loba": [("200", "13")],
            "san onofre": [("1015", "70")],
            "san pablo": [("201", "13"), ("817", "52")],
            "san pedro": [("1016", "70"), ("107", "05"), ("1099", "76")],
            "san pedro de uraba": [("108", "05")],
            "san pelayo": [("520", "23")],
            "san rafael": [("109", "05")],
            "san roque": [("110", "05")],
            "san sebastian": [("427", "19")],
            "san sebastian de buenavista": [("726", "47")],
            "san vicente": [("111", "05")],
            "san vicente del caguan": [("375", "18")],
            "santa ana": [("728", "47")],
            "santa barbara": [("112", "05"), ("819", "52"), ("982", "68")],
            "santa catalina": [("202", "13")],
            "santa maria": [("301", "15"), ("680", "41")],
            "santa marta": [("704", "47")],
            "santa rosa": [("203", "13"), ("429", "19")],
            "santa rosa de cabal": [("904", "66")],
            "santa rosa de osos": [("113", "05")],
            "santafe de antioquia": [("022", "05")],
            "santander de quilichao": [("428", "19")],
            "santo tomas": [("162", "08")],
            "santuario": [("115", "05"), ("905", "66")],
            "saravena": [("142", "81")],
            "segovia": [("116", "05")],
            "since": [("1017", "70")],
            "sincelejo": [("995", "70")],
            "soacha": [("609", "25")],
            "socorro": [("985", "68")],
            "sogamoso": [("310", "15")],
            "soledad": [("163", "08")],
            "sonson": [("117", "05")],
            "sopetran": [("118", "05")],
            "sopo": [("610", "25")],
            "sotaquira": [("313", "15")],
            "suaita": [("986", "68")],
            "suan": [("164", "08")],
            "sucre": [("1018", "70"), ("433", "19"), ("987", "68")],
            "supia": [("359", "17")],
            "tado": [("492", "27")],
            "tame": [("143", "81")],
            "tamesis": [("119", "05")],
            "taraza": [("120", "05")],
            "tarqui": [("682", "41")],
            "tauramena": [("395", "85")],
            "tibu": [("863", "54")],
            "tierralta": [("521", "23")],
            "timbiqui": [("435", "19")],
            "titiringui": [("122", "05")],
            "titiribi": [("122", "05")],
            "toca": [("324", "15")],
            "tolú viejo": [("1020", "70")],
            "tolu viejo": [("1020", "70")],
            "turbo": [("124", "05")],
            "turbaco": [("209", "13")],
            "turbana": [("210", "13")],
            "turmeque": [("329", "15")],
            "tunja": [("213", "15")],
            "tumaco": [("824", "52")],
            "ubala": [("625", "25")],
            "unguia": [("493", "27")],
            "uramita": [("125", "05")],
            "uribia": [("701", "44")],
            "urrao": [("126", "05")],
            "valdivia": [("127", "05")],
            "valencia": [("522", "23")],
            "valledupar": [("439", "20")],
            "vegachi": [("129", "05")],
            "venecia": [("130", "05"), ("586", "25")],
            "venezuela": [("1125", "000")],
            "viani": [("631", "25")],
            "vigia del fuerte": [("131", "05")],
            "villa del rosario": [("866", "54")],
            "villavicencio": [("734", "50")],
            "vista hermosa": [("762", "50")],
            "viterbo": [("362", "17")],
            "yacopi": [("636", "25")],
            "yali": [("132", "05")],
            "yarumal": [("133", "05")],
            "yolombo": [("134", "05")],
            "yopal": [("379", "85")],
            "yumbo": [("1108", "76")],
            "zambrano": [("212", "13")],
            "zaragoza": [("136", "05")],
            "zarzal": [("1109", "76")],
            "zipaquira": [("638", "25")],
        }

        matches = []

        if place in municipalities:
            matches = municipalities[place]
        else:
            paddedPlace = f" {place} "

            for municipalityName, municipalityValues in sorted(municipalities.items(), key=lambda item: len(item[0]), reverse=True,):
                if f" {municipalityName} " in paddedPlace:
                    matches = municipalityValues
                    break

        if matches:
            if detectedDepartment:
                for municipalityCode, departmentCode in matches:
                    if departmentCode == detectedDepartment:
                        return municipalityCode, departmentCode

            return matches[0]

        if detectedDepartment:
            return "000", detectedDepartment

        return "000", "000"

    def _mapAcademicLevel(self, value: str) -> str:
        academicLevel = self._normalize(value)

        if not academicLevel:
            return "000"

        if academicLevel == "primaria":
            return "001"

        if academicLevel in {"bachiller", "bachiller - 002"}:
            return "002"

        if academicLevel in {"tecnico", "tecnico - 003"}:
            return "003"

        if academicLevel in {"tecnologo", "tecnologo - 004"}:
            return "004"

        if academicLevel in {"universitario", "universitario - 005", "profesional"}:
            return "005"

        if academicLevel in {"especialista", "especializacion"}:
            return "006"

        return "000"

    def _mapContributor(self, value: str) -> str:
        contributor = self._normalize(value)

        if not contributor:
            return "1"

        if "aprendiz" in contributor and "lectiva" in contributor:
            return "12"

        if "aprendiz" in contributor and "productiva" in contributor:
            return "19"

        return "1"

    def _mapSaturdayWorkByCharge(self, value: str) -> str:
        originalValue = " ".join(str(value or "").strip().split())

        if not originalValue:
            return "S"

        charge = self._normalize(originalValue)
        chargeCode = self._normalizeCode(originalValue, 4)

        if chargeCode.isdigit():
            mappedCharge = self._mapTypeCharge(originalValue)
            charge = self._normalize(mappedCharge)

        if not charge:
            return "S"

        administrativeKeywords = [
            "aprendiz lectiva",
            "aprendiz productiva",
            "aprendiz",
            "administrativo",
            "administrativa",
            "analista",
            "auxiliar administrativo",
            "auxiliar administrativa",
            "coordinador",
            "coordinadora",
            "director",
            "directora",
            "gerente",
            "jefe",
            "lider",
            "tesoreria",
            "contabilidad",
            "contable",
            "nomina",
            "compras",
            "desarrollo humano",
            "tecnologia",
            "informatica",
            "calidad",
            "sst",
            "seguridad y salud",
            "facturacion",
            "costos",
            "bienestar",
            "legal",
            "financiero",
            "financiera",
            "impuestos",
            "mercadeo",
            "call center",
            "centro de experiencia",
        ]

        if any(keyword in charge for keyword in administrativeKeywords):
            return "N"

        return "S"

    def _mapEmployeeSalaryType(self, value: str) -> str:
        contributor = self._normalize(value)

        if not contributor:
            return "F"

        if "aprendiz" in contributor and "lectiva" in contributor:
            return "V"

        return "F"

    def _writeMasterChangesSheet(self, templateWorkbook, currentContent: bytes, previousContent: bytes | None) -> None:
        sheetName = "Cambios de Maestro"

        if sheetName not in templateWorkbook.sheetnames:
            raise ValueError(f"No se encontró la hoja {sheetName} en la plantilla.")

        worksheet = templateWorkbook[sheetName]
        self._clearMasterChangesSheet(worksheet)

        if not previousContent:
            return

        currentWorkbook = load_workbook(BytesIO(currentContent), data_only=True, keep_links=False)
        previousWorkbook = load_workbook(BytesIO(previousContent), data_only=True, keep_links=False)
        currentWorksheet = currentWorkbook.active
        previousWorksheet = previousWorkbook.active

        requiredColumns = {
            self._normalize(self.COMPANY_RUT_COLUMN),
            self._normalize("colaborador - numero de documento"),
            self._normalize("colaborador - nombre completo"),
            self._normalize("trabajo - tasa de retención (%)"),
            self._normalize("plan - fondo de pensiones"),
            self._normalize("colaborador - banco"),
            self._normalize("campos personalizados de colaborador - barrio"),
            self._normalize("trabajo - sobreescribir caja"),
            self._normalize("trabajo - cargo"),
            self._normalize("trabajo - nombre sub-area asignada(o)"),
            self._normalize("trabajo - nombre sub-area asignada(o)"),
            self._normalize("trabajo - nombre sub-area nivel 1"),
            self._normalize("trabajo - nombre sub-area asignada(o)"),
            self._normalize("trabajo - nombre sub-area asignada(o)"),
            self._normalize("campos personalizados de trabajo - dia de descanso 2 (t)"),
            self._normalize("campos personalizados de trabajo - dia de descanso 2 (t)"),
            self._normalize("colaborador - municipio"),
            self._normalize("colaborador - email"),
            self._normalize("colaborador - numero de cuenta"),
            self._normalize("colaborador - direccion"),
            self._normalize("colaborador - departamento"),
            self._normalize("plan - eps"),
            self._normalize("colaborador - estado civil"),
            self._normalize("trabajo - fecha ingreso compania"),
            self._normalize("colaborador - fecha de nacimiento"),
            self._normalize("trabajo - fecha termino trabajo"),
            self._normalize("campos personalizados de colaborador - lugar de nacimiento"),
            self._normalize("campos personalizados de colaborador - nivel academico"),
            self._normalize("colaborador - sexo"),
            self._normalize("colaborador - telefono particular"),
            self._normalize("trabajo - tipo salario"),
        }

        currentHeaderRow, currentColumnMap = self._findColumnsByNames( worksheet=currentWorksheet, requiredColumns=requiredColumns)
        previousHeaderRow, previousColumnMap = self._findColumnsByNames(worksheet=previousWorksheet, requiredColumns=requiredColumns)
        currentRows = self._getSourceRowsByCompanyRut(worksheet=currentWorksheet, headerRowNumber=currentHeaderRow, columnMap=currentColumnMap)
        previousRows = self._getSourceRowsByCompanyRut(worksheet=previousWorksheet, headerRowNumber=previousHeaderRow, columnMap=previousColumnMap)
        previousRowsByEmployee = self._buildRowsByEmployeeDocument(previousRows)

        changeFields = [
            {
                "sourceColumn": "trabajo - tasa de retención (%)",
                "fieldName": "% RETENCION",
            },
            {
                "sourceColumn": "plan - fondo de pensiones",
                "fieldName": "AFP",
                "transform": self._mapPensionFund,
            },
            {
                "sourceColumn": "colaborador - banco",
                "fieldName": "BANCO",
                "transform": self._mapBankCorporation,
            },
            {
                "sourceColumn": "campos personalizados de colaborador - barrio",
                "fieldName": "BARRIO",
            },
            {
                "sourceColumn": "trabajo - sobreescribir caja",
                "fieldName": "CAJA COMP",
                "fixedValue": "CCF04",
            },
            {
                "sourceColumn": "trabajo - cargo",
                "fieldName": "CARGO",
            },
            {
                "sourceColumn": "trabajo - nombre sub-area asignada(o)",
                "fieldName": "CEN1",
                "transform": lambda value: self._mapCostCenter1(self._mapCostCenter2(value)),
            },
            {
                "sourceColumn": "trabajo - nombre sub-area asignada(o)",
                "fieldName": "CEN2",
                "transform": self._mapCostCenter2_1,
            },
            {
                "sourceColumn": "trabajo - nombre sub-area asignada(o)",
                "fieldName": "DESC CEN2",
            },
            {
                "sourceColumn": "trabajo - nombre sub-area asignada(o)",
                "fieldName": "CEN3",
                "transform": self._mapCostCenter3,
            },
            {
                "sourceColumn": "trabajo - nombre sub-area asignada(o)",
                "fieldName": "DESC CEN3",
            },
            {
                "sourceColumn": "campos personalizados de trabajo - dia de descanso 2 (t)",
                "fieldName": "CEN4",
                "transform": self._mapRestDayCode,
            },
            {
                "sourceColumn": "campos personalizados de trabajo - dia de descanso 2 (t)",
                "fieldName": "DESC CEN4",
            },
            {
                "sourceColumn": "colaborador - municipio",
                "fieldName": "CIUDAD RESID",
                "transform": self._mapMunicipality,
            },
            {
                "sourceColumn": "colaborador - email",
                "fieldName": "CORREO ELECTRONICO",
            },
            {
                "sourceColumn": "colaborador - numero de cuenta",
                "fieldName": "CUENTA",
            },
            {
                "sourceColumn": "colaborador - direccion",
                "fieldName": "DIRECCION",
            },
            {
                "sourceColumn": "colaborador - departamento",
                "fieldName": "DPTO RESID",
                "transform": self._mapDepartment,
            },
            {
                "sourceColumn": "plan - eps",
                "fieldName": "EPS",
            },
            {
                "sourceColumn": "colaborador - estado civil",
                "fieldName": "ESTADO CIVIL",
                "transform": self._mapMaritalStatus,
            },
            {
                "sourceColumn": "trabajo - fecha ingreso compania",
                "fieldName": "FINGRESO",
                "transform": self._formatDateAsText,
            },
            {
                "sourceColumn": "colaborador - fecha de nacimiento",
                "fieldName": "FNACIMIEN",
                "transform": self._formatDateAsText,
            },
            {
                "sourceColumn": "trabajo - fecha termino trabajo",
                "fieldName": "FRETIRO",
                "transform": self._formatDateAsText,
            },
            {
                "sourceColumn": "campos personalizados de colaborador - lugar de nacimiento",
                "fieldName": "LUGAR NACIM",
                "transform": self._mapMunicipality,
            },
            {
                "sourceColumn": "campos personalizados de colaborador - nivel academico",
                "fieldName": "NIVEL EST.",
                "transform": self._mapAcademicLevel,
            },
            {
                "sourceColumn": "colaborador - sexo",
                "fieldName": "SEXO",
            },
            {
                "sourceColumn": "colaborador - telefono particular",
                "fieldName": "TELEFONO",
            },
            {
                "sourceColumn": "trabajo - tipo salario",
                "fieldName": "TIPO CONTRATO",
                "fixedValue": "fijo",
            },
        ]

        currentRowNumber = 2
        currentDateText = datetime.now().strftime("%d/%m/%Y")
        currentEmployeeOccurrences = {}

        for currentRow in currentRows:
            employeeDocument = self._cleanDocument(self._getCellText(currentRow[self._normalize("colaborador - numero de documento")]))

            if not employeeDocument:
                continue

            occurrenceIndex = currentEmployeeOccurrences.get(employeeDocument, 0)
            currentEmployeeOccurrences[employeeDocument] = occurrenceIndex + 1
            previousEmployeeRows = previousRowsByEmployee.get(employeeDocument, [])

            if occurrenceIndex >= len(previousEmployeeRows):
                continue

            previousRow = previousEmployeeRows[occurrenceIndex]
            employeeName = self._getCellText(currentRow[self._normalize("colaborador - nombre completo")]).upper()

            for changeField in changeFields:
                sourceColumn = self._normalize(changeField["sourceColumn"])
                fieldName = changeField["fieldName"]
                previousValue = self._getCellText(previousRow.get(sourceColumn))
                currentValue = self._getCellText(currentRow.get(sourceColumn))

                if self._normalize(previousValue) == self._normalize(currentValue):
                    continue

                fixedValue = changeField.get("fixedValue")
                transform = changeField.get("transform")

                if fixedValue is not None:
                    newValue = fixedValue
                elif transform:
                    newValue = transform(currentValue)
                else:
                    newValue = currentValue

                worksheet.cell(row=currentRowNumber, column=1, value=employeeDocument)
                worksheet.cell(row=currentRowNumber, column=2, value=employeeName)
                worksheet.cell(row=currentRowNumber, column=3, value=fieldName)
                worksheet.cell(row=currentRowNumber, column=4, value=newValue)
                worksheet.cell(row=currentRowNumber, column=5, value=currentDateText)

                for columnNumber in range(1, 6):
                    worksheet.cell(row=currentRowNumber, column=columnNumber).number_format = "@"

                currentRowNumber += 1

    def _clearMasterChangesSheet(self, worksheet) -> None:
        startRow = 2

        for rowNumber in range(startRow, worksheet.max_row + 1):
            for columnNumber in range(1, 6):
                worksheet.cell(row=rowNumber, column=columnNumber, value=None)

    def _findColumnsByNames(self, worksheet, requiredColumns: set[str]):
        for row in worksheet.iter_rows(min_row=1, max_row=10):
            currentColumns = {}

            for cell in row:
                normalizedValue = self._normalize(cell.value)

                if normalizedValue in requiredColumns:
                    currentColumns[normalizedValue] = cell.column

            if all(columnName in currentColumns for columnName in requiredColumns):
                return row[0].row, currentColumns

        missingColumnsText = ", ".join(sorted(requiredColumns))
        raise ValueError(f"No se encontraron las columnas requeridas para Cambios de Maestro: {missingColumnsText}.")

    def _getSourceRowsByCompanyRut(self, worksheet, headerRowNumber: int, columnMap: dict) -> list[dict[str, Any]]:
        sourceRows = []

        companyRutColumn = columnMap[self._normalize(self.COMPANY_RUT_COLUMN)]
        expectedCompanyRut = self._cleanRut(self.COMPANY_RUT_VALUE)

        for rowNumber in range(headerRowNumber + 1, worksheet.max_row + 1):
            companyRutValue = worksheet.cell(row=rowNumber, column=companyRutColumn).value
            currentCompanyRut = self._cleanRut(companyRutValue)

            if currentCompanyRut != expectedCompanyRut:
                continue

            sourceRows.append(self._buildSourceRowContext(worksheet=worksheet, rowNumber=rowNumber, columnMap=columnMap,))

        return sourceRows

    def _buildRowsByEmployeeDocument(self, sourceRows: list[dict[str, Any]]) -> dict[str, list[dict[str, Any]]]:
        rowsByEmployee = {}
        documentColumn = self._normalize("colaborador - numero de documento")

        for sourceRow in sourceRows:
            employeeDocument = self._cleanDocument(self._getCellText(sourceRow.get(documentColumn)))

            if not employeeDocument:
                continue

            if employeeDocument not in rowsByEmployee:
                rowsByEmployee[employeeDocument] = []

            rowsByEmployee[employeeDocument].append(sourceRow)

        return rowsByEmployee

    def _cleanDocument(self, value) -> str:
        return "".join(character for character in str(value or "") if character.isdigit())

    def _writeRetirementsSheet(self, templateWorkbook, currentContent: bytes, dateFrom: date, dateTo: date) -> None:
        sheetName = "Retiros"

        if sheetName not in templateWorkbook.sheetnames:
            raise ValueError(f"No se encontró la hoja {sheetName} en la plantilla.")

        worksheet = templateWorkbook[sheetName]
        self._clearRetirementsSheet(worksheet)

        currentWorkbook = load_workbook(BytesIO(currentContent), data_only=True, keep_links=False)
        currentWorksheet = currentWorkbook.active

        requiredColumns = {
            self._normalize(self.COMPANY_RUT_COLUMN),
            self._normalize("colaborador - numero de documento"),
            self._normalize("colaborador - nombre completo"),
            self._normalize("trabajo - fecha termino trabajo"),
            self._normalize("trabajo - razon de termino"),
        }

        headerRow, columnMap = self._findColumnsByNames(worksheet=currentWorksheet, requiredColumns=requiredColumns)
        sourceRows = self._getSourceRowsByCompanyRut(worksheet=currentWorksheet, headerRowNumber=headerRow, columnMap=columnMap)

        documentColumn = self._normalize("colaborador - numero de documento")
        nameColumn = self._normalize("colaborador - nombre completo")
        retirementDateColumn = self._normalize("trabajo - fecha termino trabajo")
        reasonColumn = self._normalize("trabajo - razon de termino")

        currentRowNumber = 3

        for sourceRow in sourceRows:
            retirementDateCell = sourceRow.get(retirementDateColumn)

            retirementDateValue = (
                retirementDateCell.value
                if hasattr(retirementDateCell, "value")
                else retirementDateCell
            )

            retirementDate = self._parseDate(retirementDateValue)

            if not retirementDate:
                continue

            if retirementDate < dateFrom or retirementDate > dateTo:
                continue

            employeeDocument = self._cleanDocument(self._getCellText(sourceRow.get(documentColumn)))
            employeeName = self._getCellText(sourceRow.get(nameColumn)).upper()
            retirementDateText = self._formatDateAsText(sourceRow.get(retirementDateColumn))
            reason = self._getCellText(sourceRow.get(reasonColumn))

            worksheet.cell(row=currentRowNumber, column=1, value=employeeDocument)
            worksheet.cell(row=currentRowNumber, column=2, value=employeeName)
            worksheet.cell(row=currentRowNumber, column=3, value=retirementDateText)
            worksheet.cell(row=currentRowNumber, column=4, value=reason)

            for columnNumber in range(1, 5):
                worksheet.cell(row=currentRowNumber, column=columnNumber).number_format = "@"

            currentRowNumber += 1

    def _clearRetirementsSheet(self, worksheet) -> None:
        startRow = 3

        for rowNumber in range(startRow, worksheet.max_row + 1):
            for columnNumber in range(1, 5):
                worksheet.cell(row=rowNumber, column=columnNumber).value = None


    def _writeSalaryChangesSheet(self, templateWorkbook, currentContent: bytes, previousContent: bytes | None) -> None:
        sheetName = "Modificaciones de Salario"

        if sheetName not in templateWorkbook.sheetnames:
            raise ValueError(f"No se encontró la hoja {sheetName} en la plantilla.")

        worksheet = templateWorkbook[sheetName]
        self._clearSalaryChangesSheet(worksheet)

        if not previousContent:
            return

        currentWorkbook = load_workbook(BytesIO(currentContent), data_only=True, keep_links=False)
        previousWorkbook = load_workbook(BytesIO(previousContent), data_only=True, keep_links=False)
        currentWorksheet = currentWorkbook.active
        previousWorksheet = previousWorkbook.active

        requiredColumns = {
            self._normalize(self.COMPANY_RUT_COLUMN),
            self._normalize("colaborador - numero de documento"),
            self._normalize("colaborador - nombre completo"),
            self._normalize("trabajo - fecha ingreso compania"),
            self._normalize("campos personalizados de trabajo - salario (obligatorio)"),
        }

        currentHeaderRow, currentColumnMap = self._findColumnsByNames(worksheet=currentWorksheet, requiredColumns=requiredColumns)
        previousHeaderRow, previousColumnMap = self._findColumnsByNames(worksheet=previousWorksheet, requiredColumns=requiredColumns)
        currentRows = self._getSourceRowsByCompanyRut(worksheet=currentWorksheet, headerRowNumber=currentHeaderRow, columnMap=currentColumnMap)
        previousRows = self._getSourceRowsByCompanyRut(worksheet=previousWorksheet, headerRowNumber=previousHeaderRow, columnMap=previousColumnMap)

        previousRowsByEmployee = self._buildRowsByEmployeeDocument(previousRows)

        documentColumn = self._normalize("colaborador - numero de documento")
        nameColumn = self._normalize("colaborador - nombre completo")
        incomeDateColumn = self._normalize("trabajo - fecha ingreso compania")
        salaryColumn = self._normalize("campos personalizados de trabajo - salario (obligatorio)")

        currentRowNumber = 2
        currentDateText = datetime.now().strftime("%d/%m/%Y")
        currentEmployeeOccurrences = {}

        for currentRow in currentRows:
            employeeDocument = self._cleanDocument(self._getCellText(currentRow.get(documentColumn)))

            if not employeeDocument:
                continue

            occurrenceIndex = currentEmployeeOccurrences.get(employeeDocument, 0)
            currentEmployeeOccurrences[employeeDocument] = occurrenceIndex + 1
            previousEmployeeRows = previousRowsByEmployee.get(employeeDocument, [])

            if occurrenceIndex >= len(previousEmployeeRows):
                continue

            previousRow = previousEmployeeRows[occurrenceIndex]
            previousSalary = self._getCellText(previousRow.get(salaryColumn))
            currentSalary = self._getCellText(currentRow.get(salaryColumn))

            if self._normalizeSalaryForCompare(previousSalary) == self._normalizeSalaryForCompare(currentSalary):
                continue

            employeeName = self._getCellText(currentRow.get(nameColumn)).upper()
            incomeDateText = self._formatDateAsText(currentRow.get(incomeDateColumn))

            worksheet.cell(row=currentRowNumber, column=1, value=employeeDocument)
            worksheet.cell(row=currentRowNumber, column=2, value=employeeName)
            worksheet.cell(row=currentRowNumber, column=3, value=incomeDateText)
            worksheet.cell(row=currentRowNumber, column=4, value=currentDateText)
            worksheet.cell(row=currentRowNumber, column=5, value=previousSalary)
            worksheet.cell(row=currentRowNumber, column=6, value=currentSalary)

            for columnNumber in range(1, 7):
                worksheet.cell(row=currentRowNumber, column=columnNumber).number_format = "@"

            currentRowNumber += 1
    def _clearSalaryChangesSheet(self, worksheet) -> None:
        startRow = 2

        for rowNumber in range(startRow, worksheet.max_row + 1):
            for columnNumber in range(1, 10):
                worksheet.cell(row=rowNumber, column=columnNumber).value = None
    
    def _normalizeSalaryForCompare(self, value) -> str:
        text = self._getCellText(value)

        if not text:
            return ""

        text = text.strip()
        text = text.replace("$", "")
        text = text.replace(" ", "")
        text = text.replace(".", "")
        text = text.replace(",", ".")

        try:
            numberValue = float(text)

            if numberValue.is_integer():
                return str(int(numberValue))

            return str(numberValue)
        except ValueError:
            return self._normalize(text)